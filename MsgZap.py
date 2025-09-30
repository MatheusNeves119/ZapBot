#importando bibliotecas
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait # <<< IMPORTAÇÃO PARA ESPERA INTELIGENTE
from selenium.webdriver.support import expected_conditions as EC # <<< IMPORTAÇÃO PARA ESPERA INTELIGENTE
import pyautogui as tempoEspera
from openpyxl import load_workbook
import random
import pyperclip

print("--- CHECKPOINT 1: Script iniciado ---")

# --- CAMINHOS DOS ARQUIVOS ---
nome_arquivo_contatos = r"C:\Mensagem automatica Zap\Contatos.xlsx"
imagem_para_enviar = r"C:\Mensagem automatica Zap\Foto_Tenis.jpeg"

try:
    planilhaDadosContato = load_workbook(nome_arquivo_contatos)
    # !!! ATENÇÃO: Verifique se o nome da sua aba é 'Dados' ou 'Planilha1' e ajuste aqui !!!
    sheet_selecionada = planilhaDadosContato['Dados'] 
    print("--- CHECKPOINT 2: Arquivo Excel carregado com sucesso! ---")
except FileNotFoundError:
    print(f"\nERRO GRAVE: Arquivo Excel não encontrado em '{nome_arquivo_contatos}'")
    input("Pressione Enter para fechar.")
    exit()
except KeyError:
    print("\nERRO GRAVE: A aba da planilha não foi encontrada! Verifique o nome ('Dados', 'Planilha1', etc.).")
    input("Pressione Enter para fechar.")
    exit()

# --- INICIALIZAÇÃO DO NAVEGADOR ---
navegadorChrome = webdriver.Chrome()
print("--- CHECKPOINT 3: Navegador aberto ---")

navegadorChrome.get("https://web.whatsapp.com")
print("--- CHECKPOINT 4: Aguardando login (escaneie o QR Code) ---")

while len(navegadorChrome.find_elements(By.ID, 'side')) < 1:
    tempoEspera.sleep(1)
print("--- CHECKPOINT 5: Login detectado! ---")
tempoEspera.sleep(random.uniform(3, 5))

# --- LOOP PRINCIPAL PARA ENVIO DAS MENSAGENS ---
print("--- CHECKPOINT 6: Iniciando loop para ler contatos... ---")
if sheet_selecionada.max_row < 2:
    print("AVISO: A planilha não tem contatos para enviar.")
else:
    for linha_num in range(2 , sheet_selecionada.max_row + 1):
        nomeContato = sheet_selecionada[f'A{linha_num}'].value
        mensagemContato = sheet_selecionada[f'B{linha_num}'].value
        
        print(f"--- Lendo linha {linha_num-1}: Contato '{nomeContato}' ---")
        try:
            # --- SOLUÇÃO DE ESTABILIDADE: ESPERA INTELIGENTE PELA BARRA DE PESQUISA ---
            # O robô vai esperar até 10 segundos a barra de pesquisa estar pronta antes de continuar
            wait = WebDriverWait(navegadorChrome, 10)
            caixa_de_pesquisa = wait.until(
                EC.presence_of_element_located((By.XPATH, '//div[@aria-placeholder="Pesquisar ou começar uma nova conversa"]'))
            )
            
            # PARTE 1: ABRIR A CONVERSA
            caixa_de_pesquisa.clear()
            caixa_de_pesquisa.send_keys(nomeContato)
            tempoEspera.sleep(random.uniform(2, 4))
            caixa_de_pesquisa.send_keys(Keys.ENTER)
            tempoEspera.sleep(random.uniform(3, 5))

            # PARTE 2: ANEXAR A IMAGEM
            botao_mais = navegadorChrome.find_element(By.CSS_SELECTOR, "span[data-icon='plus-rounded']")
            botao_mais.click()
            tempoEspera.sleep(random.uniform(2, 3))

            input_de_arquivo = navegadorChrome.find_element(By.XPATH, '//input[@accept="image/*,video/mp4,video/3gpp,video/quicktime"]')
            input_de_arquivo.send_keys(imagem_para_enviar)
            tempoEspera.sleep(random.uniform(3, 5))

            # PARTE 3: ENVIAR A IMAGEM COM A LEGENDA (LÓGICA CORRETA)
            saudacoes = ["Olá!", "Bom dia!", "Boa tarde!", "Confira a novidade:", "Chegou na loja:"]
            saudacao_aleatoria = random.choice(saudacoes)
            mensagem_final = f"{saudacao_aleatoria}\n\n{mensagemContato}"
            pyperclip.copy(mensagem_final)

            caixa_de_legenda = navegadorChrome.find_element(By.XPATH, '//*[@id="app"]//div[@contenteditable="true"]')
            caixa_de_legenda.send_keys(Keys.CONTROL, "v")
            tempoEspera.sleep(random.uniform(2, 4))

            botao_enviar = navegadorChrome.find_element(By.XPATH, '//div[@aria-label="Enviar"]')
            botao_enviar.click()
            
            print(f"Imagem e mensagem enviadas para '{nomeContato}'!")

            tempo_de_espera = random.uniform(3, 7)
            print(f"Aguardando {tempo_de_espera:.2f} segundos antes do próximo envio...")
            tempoEspera.sleep(tempo_de_espera)
            
        except Exception as e:
            print(f"Erro ao processar o contato '{nomeContato}': {e}")

print("--- CHECKPOINT 7: Fim do script. ---")
input("Pressione Enter para fechar.")
navegadorChrome.quit()